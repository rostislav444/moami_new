from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.conf import settings

import logging
import json
import io

logger = logging.getLogger(__name__)


class AIAssistantViewSet(viewsets.ViewSet):
    """
    AI Ассистент для автоматического маппинга товаров

    Endpoints:
    - POST /api/ai/suggest-category/ - предложить категорию для товара
    - POST /api/ai/suggest-attributes/ - предложить атрибуты для товара
    - POST /api/ai/auto-map-product/ - полный автомаппинг товара
    - POST /api/ai/auto-map-bulk/ - массовый автомаппинг товаров
    - POST /api/ai/parse_attributes_file/ - парсинг XLSX/CSV файла с атрибутами
    """
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def _get_agent(self, marketplace_id: int):
        """Получить AI агента для маркетплейса"""
        from apps.marketplaces.models import Marketplace
        from apps.marketplaces.services.ai_mapping_agent import AIMarketplaceMappingAgent

        try:
            marketplace = Marketplace.objects.get(id=marketplace_id)
            return AIMarketplaceMappingAgent(marketplace)
        except Marketplace.DoesNotExist:
            return None
        except ImportError as e:
            logger.error(f"Failed to import AI agent: {e}")
            return None

    @action(detail=False, methods=['post'])
    def suggest_category(self, request):
        """
        Предложить категорию маркетплейса для товара

        POST /api/ai/suggest-category/
        Body: {
            "marketplace_id": 1,
            "product_id": 123
        }
        """
        from apps.product.models import Product

        marketplace_id = request.data.get('marketplace_id')
        product_id = request.data.get('product_id')

        if not marketplace_id or not product_id:
            return Response(
                {'error': 'marketplace_id and product_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response(
                {'error': 'Product not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        agent = self._get_agent(marketplace_id)
        if not agent:
            return Response(
                {'error': 'AI agent not available. Check ANTHROPIC_API_KEY setting.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

        try:
            result = agent.suggest_category(product)
            return Response({
                'success': True,
                'suggestion': result
            })
        except Exception as e:
            logger.error(f"AI suggest_category error: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def suggest_attributes(self, request):
        """
        Предложить атрибуты для товара

        POST /api/ai/suggest-attributes/
        Body: {
            "marketplace_id": 1,
            "product_id": 123,
            "category_code": "6390",
            "include_optional": false
        }
        """
        from apps.product.models import Product

        marketplace_id = request.data.get('marketplace_id')
        product_id = request.data.get('product_id')
        category_code = request.data.get('category_code')
        include_optional = request.data.get('include_optional', False)

        if not all([marketplace_id, product_id, category_code]):
            return Response(
                {'error': 'marketplace_id, product_id, and category_code are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response(
                {'error': 'Product not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        agent = self._get_agent(marketplace_id)
        if not agent:
            return Response(
                {'error': 'AI agent not available'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

        try:
            result = agent.suggest_attributes(product, category_code, include_optional)
            return Response({
                'success': True,
                'attributes': result
            })
        except Exception as e:
            logger.error(f"AI suggest_attributes error: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def auto_map_product(self, request):
        """
        Полный автоматический маппинг товара

        POST /api/ai/auto-map-product/
        Body: {
            "marketplace_id": 1,
            "product_id": 123
        }
        """
        from apps.product.models import Product

        marketplace_id = request.data.get('marketplace_id')
        product_id = request.data.get('product_id')

        if not marketplace_id or not product_id:
            return Response(
                {'error': 'marketplace_id and product_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response(
                {'error': 'Product not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        agent = self._get_agent(marketplace_id)
        if not agent:
            return Response(
                {'error': 'AI agent not available'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

        try:
            result = agent.auto_map_product(product)
            return Response(result)
        except Exception as e:
            logger.error(f"AI auto_map_product error: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def auto_map_bulk(self, request):
        """
        Массовый автомаппинг товаров

        POST /api/ai/auto-map-bulk/
        Body: {
            "marketplace_id": 1,
            "product_ids": [1, 2, 3],
            "category_id": 5  // опционально - фильтр по категории
        }
        """
        from apps.product.models import Product

        marketplace_id = request.data.get('marketplace_id')
        product_ids = request.data.get('product_ids', [])
        category_id = request.data.get('category_id')

        if not marketplace_id:
            return Response(
                {'error': 'marketplace_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        agent = self._get_agent(marketplace_id)
        if not agent:
            return Response(
                {'error': 'AI agent not available'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

        # Получить товары
        if product_ids:
            products = Product.objects.filter(id__in=product_ids)
        elif category_id:
            products = Product.objects.filter(category_id=category_id)[:50]  # Лимит
        else:
            products = Product.objects.all()[:20]  # Лимит

        results = []
        for product in products:
            try:
                result = agent.auto_map_product(product)
                results.append(result)
            except Exception as e:
                results.append({
                    'product_id': product.id,
                    'error': str(e)
                })

        return Response({
            'success': True,
            'processed': len(results),
            'results': results
        })

    @action(detail=False, methods=['get'])
    def status(self, request):
        """
        Проверить статус AI ассистента

        GET /api/ai/status/
        """
        has_api_key = bool(getattr(settings, 'ANTHROPIC_API_KEY', None))

        try:
            import anthropic
            has_library = True
        except ImportError:
            has_library = False

        return Response({
            'available': has_api_key and has_library,
            'has_api_key': has_api_key,
            'has_library': has_library,
            'message': 'AI assistant is ready' if (has_api_key and has_library) else 'Missing API key or anthropic library'
        })

    @action(detail=False, methods=['post'])
    def parse_attributes_file(self, request):
        """
        Универсальный парсинг XLSX/CSV файла с атрибутами.

        Процесс:
        1. Читаем файл, собираем список вкладок с превью
        2. AI определяет какие вкладки содержат данные (характеристики, размеры и т.д.)
        3. Цикл: для каждой полезной вкладки — AI определяет структуру, Python парсит
        4. Конкатенируем результаты и отдаём

        POST /api/ai/parse_attributes_file/
        Content-Type: multipart/form-data
        Body: file (XLSX или CSV)
        """
        import anthropic
        import openpyxl
        import csv
        import re as _re

        if 'file' not in request.FILES:
            return Response({'error': 'Файл не загружен'}, status=400)

        uploaded_file = request.FILES['file']
        filename = uploaded_file.name.lower()
        file_bytes = uploaded_file.read()
        marketplace_id = request.data.get('marketplace_id') or request.query_params.get('marketplace_id')

        # ===== ШАГ 1: Читаем все вкладки =====
        all_sheets = {}  # {name: [[row], ...]}
        try:
            if filename.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                for sn in wb.sheetnames:
                    ws = wb[sn]
                    sheet_rows = []
                    for row in ws.iter_rows(values_only=True):
                        if any(c is not None for c in row):
                            sheet_rows.append([str(c).strip() if c else '' for c in row])
                    all_sheets[sn] = sheet_rows
                wb.close()
            elif filename.endswith('.csv'):
                content = file_bytes.decode('utf-8-sig')
                sep = '\t' if '\t' in content else ','
                reader = csv.reader(io.StringIO(content), delimiter=sep)
                all_sheets['CSV'] = [r for r in reader if any(c.strip() for c in r)]
            else:
                return Response({'error': 'Формат не поддерживается'}, status=400)
        except Exception as e:
            return Response({'error': f'Ошибка чтения: {e}'}, status=400)

        if not all_sheets:
            return Response({'error': 'Файл пустой'}, status=400)

        sheet_names = list(all_sheets.keys())
        print(f"[PARSE] XLSX sheets: {sheet_names}")

        # ===== ШАГ 2: AI определяет какие вкладки парсить =====
        # Если одна вкладка — не спрашиваем AI
        if len(sheet_names) == 1:
            sheets_to_parse = [{'sheet': sheet_names[0], 'type': 'attributes'}]
        else:
            # Превью каждой вкладки для AI
            preview = ""
            for sn, rows in all_sheets.items():
                preview += f'\n=== "{sn}" ({len(rows)} строк) ===\n'
                for r in rows[:5]:
                    preview += '\t'.join(r[:12]) + '\n'

            try:
                client = anthropic.Anthropic(api_key=getattr(settings, 'ANTHROPIC_API_KEY', None))
                resp = client.messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=500,
                    messages=[{'role': 'user', 'content': f"""Вот вкладки Excel файла маркетплейса:
{preview}

Определи какие вкладки содержат ПОЛЕЗНЫЕ данные. Типы:
- "attributes" — характеристики/атрибуты товаров (названия, типы, возможные значения). Выбери ОДНУ вкладку (укр предпочтительнее)
- "size_grid" — размерные сетки (таблицы размеров с обхватами)
- "skip" — шаблоны товаров, примеры, инструкции, правила, дубли на других языках

НЕ включай дубли характеристик на разных языках — только одну вкладку!

Ответь ТОЛЬКО JSON массивом:
[{{"sheet": "Характеристики (укр)", "type": "attributes"}}, {{"sheet": "Розмірні сітки", "type": "size_grid"}}]"""}],
                )
                txt = resp.content[0].text.strip()
                if '```' in txt:
                    txt = txt.split('```json')[-1].split('```')[0] if '```json' in txt else txt.split('```')[1].split('```')[0]
                sheets_to_parse = json.loads(txt.strip())
                sheets_to_parse = [s for s in sheets_to_parse if s.get('type') != 'skip']
                print(f"[PARSE] AI sheets to parse: {sheets_to_parse}")
            except Exception as e:
                logger.warning(f"AI sheet detection failed: {e}, using all sheets")
                sheets_to_parse = [{'sheet': sn, 'type': 'attributes'} for sn in sheet_names]

        # ===== ШАГ 3: Парсим каждую вкладку =====
        all_attributes = []
        all_size_grids = []

        for sheet_info in sheets_to_parse:
            sn = sheet_info.get('sheet', '')
            stype = sheet_info.get('type', 'attributes')
            if sn not in all_sheets:
                continue
            rows = all_sheets[sn]
            if len(rows) < 2:
                continue

            print(f"[PARSE] Parsing sheet '{sn}' as {stype}")

            if stype == 'size_grid':
                # Parse size grid blocks into attributes
                # Each block = one attribute (e.g. "ТР жіночого одягу (3XS-15XL):226")
                # Options = sizes with their IDs
                current_grid_name = sn
                current_grid_id = None
                current_headers = None
                current_sizes = []

                def flush_grid():
                    if current_sizes and current_grid_name:
                        # Create attribute from size grid
                        values = []
                        for sz in current_sizes:
                            # First column = size name, second = ID
                            label = sz.get('label', '')
                            sid = sz.get('id', '')
                            if label:
                                values.append(f"{sid}:{label}" if sid else label)

                        all_attributes.append({
                            'name': current_grid_name.split(':')[0].strip() if ':' in current_grid_name else current_grid_name,
                            'code': current_grid_id or f'size_grid_{len(all_attributes)}',
                            'type': 'select',
                            'is_required': False,
                            'possible_values': [sz.get('label', '') for sz in current_sizes if sz.get('label')],
                            'unit': None,
                            '_sheet': sn,
                            '_size_grid_data': current_sizes,
                        })

                for r in rows:
                    non_empty = sum(1 for c in r if c)

                    # Detect new block header: "ТР жіночого одягу (3XS-15XL):226"
                    if non_empty <= 2 and r[0] and ':' in r[0]:
                        flush_grid()
                        current_sizes = []
                        raw_name = r[0]
                        if raw_name.rsplit(':', 1)[1].strip().isdigit():
                            current_grid_id = raw_name.rsplit(':', 1)[1].strip()
                            current_grid_name = raw_name.rsplit(':', 1)[0].strip()
                        else:
                            current_grid_name = raw_name
                            current_grid_id = None
                        current_headers = None
                        continue
                    elif non_empty <= 2 and r[0]:
                        flush_grid()
                        current_sizes = []
                        current_grid_name = r[0]
                        current_grid_id = None
                        current_headers = None
                        continue

                    # Detect column headers
                    r_lower = ' '.join(c.lower() for c in r if c)
                    if any(kw in r_lower for kw in ['id', 'міжнародна', 'українська', 'обхват', 'международная']):
                        current_headers = r
                        continue

                    # Data row
                    if current_headers and any(r):
                        size_entry = {'label': r[0] if r[0] else ''}
                        for i, h in enumerate(current_headers):
                            if h and i < len(r) and r[i]:
                                if h.strip().upper() == 'ID':
                                    size_entry['id'] = r[i]
                                else:
                                    size_entry[h] = r[i]
                        if size_entry.get('label'):
                            current_sizes.append(size_entry)

                flush_grid()
                continue

            # Parse attributes sheet
            # Detect format: columns (Kasta) or rows
            header = rows[0]
            is_columns_format = any(':' in h and h.rsplit(':', 1)[1].strip().isdigit() for h in header if h)

            if is_columns_format:
                # Columns format: each column = attribute
                for col_idx, raw_h in enumerate(header):
                    if not raw_h or not raw_h.strip():
                        continue
                    raw = raw_h.strip()
                    attr_id = None
                    name = raw
                    attr_type = 'select'
                    is_required = False
                    unit = None

                    if ':' in raw and raw.rsplit(':', 1)[1].strip().isdigit():
                        parts = raw.rsplit(':', 1)
                        name_part = parts[0].strip()
                        attr_id = parts[1].strip()

                        if '*' in name_part:
                            is_required = True
                            name_part = name_part.replace('*', '')

                        tm = _re.search(r'\(([^)]+)\)', name_part)
                        if tm:
                            hint = tm.group(1).lower()
                            if 'множ' in hint: attr_type = 'multiselect'
                            elif 'числ' in hint: attr_type = 'float'
                            elif 'строк' in hint or 'стрічк' in hint: attr_type = 'string'
                            name_part = name_part[:tm.start()].strip()

                        if ',' in name_part:
                            nm, mu = name_part.split(',', 1)
                            if len(mu.strip()) <= 5:
                                unit = mu.strip()
                                name_part = nm.strip()

                        name = name_part.strip()

                    values = sorted(set(
                        r[col_idx] for r in rows[1:] if col_idx < len(r) and r[col_idx]
                    ))[:200]

                    ftype = attr_type if attr_type in ('float', 'string') else (attr_type if values else 'string')

                    all_attributes.append({
                        'name': name, 'code': attr_id or f'col_{col_idx}',
                        'type': ftype, 'is_required': is_required,
                        'possible_values': values, 'unit': unit,
                        '_sheet': sn,
                    })
            else:
                # Rows format: AI determines columns
                try:
                    client = anthropic.Anthropic(api_key=getattr(settings, 'ANTHROPIC_API_KEY', None))
                    sample = '\n'.join('\t'.join(r) for r in rows[:20])
                    resp = client.messages.create(
                        model='claude-haiku-4-5-20251001',
                        max_tokens=500,
                        messages=[{'role': 'user', 'content': f"""Вкладка "{sn}" с атрибутами (строки):
{sample}

Определи номера колонок (0-based). Каждая строка = один вариант значения атрибута.
Одинаковый attr_id означает тот же атрибут, но другое значение (опция).

{{"header_row": 0, "data_start_row": 1, "columns": {{"attr_id": 0, "attr_name": 1, "attr_type": 2, "is_required": null, "value_id": 5, "value_name": 6, "unit": null}}}}

value_id = колонка с ID/кодом варианта значения (опции)
value_name = колонка с текстом значения (опции)
Если нет заголовка — header_row: null, data_start_row: 0
Ответь ТОЛЬКО JSON."""}],
                    )
                    txt = resp.content[0].text.strip()
                    if '```' in txt:
                        txt = txt.split('```json')[-1].split('```')[0] if '```json' in txt else txt.split('```')[1].split('```')[0]
                    cfg = json.loads(txt.strip())
                except Exception as e:
                    logger.warning(f"AI row config failed for {sn}: {e}")
                    continue

                cols = cfg.get('columns', {})
                ci = cols.get('attr_id')
                cn = cols.get('attr_name')
                ct = cols.get('attr_type')
                cr = cols.get('is_required')
                cvi = cols.get('value_id')
                cvn = cols.get('value_name')
                cu = cols.get('unit')
                if ci is None or cn is None:
                    continue

                start = cfg.get('data_start_row', 1) or 0
                # Map types from marketplace format to our format
                type_map_simple = {
                    'listvalues': 'multiselect', 'combobox': 'select',
                    'string': 'string', 'text': 'text',
                    'integer': 'int', 'float': 'float', 'number': 'float',
                    'boolean': 'boolean', 'checkbox': 'boolean',
                }

                attrs_dict = {}
                for r in rows[start:]:
                    aid = r[ci] if ci is not None and ci < len(r) else None
                    aname = r[cn] if cn is not None and cn < len(r) else None
                    if not aid or not aname:
                        continue
                    atype_raw = r[ct] if ct is not None and ct < len(r) else ''
                    areq = r[cr] if cr is not None and cr < len(r) else ''
                    aval_id = r[cvi] if cvi is not None and cvi < len(r) else ''
                    aval_name = r[cvn] if cvn is not None and cvn < len(r) else ''
                    aunit = r[cu] if cu is not None and cu < len(r) else ''

                    # Map type
                    atype = type_map_simple.get(str(atype_raw).lower().strip(), atype_raw or 'string')

                    if aid not in attrs_dict:
                        is_req = str(areq).lower().strip() in ('да', 'yes', 'true', '+', '1', 'так', 'нет')
                        # "Нет" means not required for Rozetka
                        if str(areq).lower().strip() in ('нет', 'no', 'false', '0'):
                            is_req = False
                        attrs_dict[aid] = {
                            'name': aname, 'code': str(aid), 'type': atype,
                            'is_required': is_req,
                            'possible_values': [], 'unit': aunit.strip() if aunit else None,
                            '_sheet': sn,
                            '_option_ids': [],
                        }
                    if aval_name and aval_name.strip():
                        attrs_dict[aid]['possible_values'].append(aval_name.strip())
                        if aval_id:
                            attrs_dict[aid]['_option_ids'].append(str(aval_id).strip())

                all_attributes.extend(attrs_dict.values())
                print(f"[PARSE] Rows format: {len(attrs_dict)} attrs from {sn}")

        print(f"[PARSE] Total: {len(all_attributes)} attrs, {len(all_size_grids)} size grids")

        # ===== ШАГ 4: Сохраняем в БД если marketplace_id указан =====
        saved_count = 0
        if marketplace_id:
            try:
                from apps.marketplaces.models import (
                    MarketplaceAttributeSet, MarketplaceAttribute, MarketplaceAttributeOption,
                    Marketplace,
                )
                mp = Marketplace.objects.get(id=marketplace_id)

                # Determine category code from request or first parsed
                category_code = request.data.get('category_code', '')

                # Find or create attribute set
                attr_set, _ = MarketplaceAttributeSet.objects.get_or_create(
                    marketplace=mp,
                    external_code=category_code or 'imported',
                    defaults={'name': category_code or 'Imported from XLSX'},
                )

                for attr_data in all_attributes:
                    code = str(attr_data.get('code', ''))
                    if not code:
                        continue

                    # Map types
                    atype = attr_data.get('type', 'string').lower()
                    type_map = {'float': 'float', 'int': 'int', 'string': 'string',
                                'text': 'text', 'select': 'select', 'multiselect': 'multiselect',
                                'boolean': 'boolean'}
                    mapped_type = type_map.get(atype, 'string')

                    ma, created = MarketplaceAttribute.objects.update_or_create(
                        attribute_set=attr_set,
                        external_code=code,
                        defaults={
                            'name': attr_data.get('name', ''),
                            'attr_type': mapped_type,
                            'is_required': attr_data.get('is_required', False),
                            'suffix': attr_data.get('unit', '') or '',
                        }
                    )

                    # Create options
                    if mapped_type in ('select', 'multiselect'):
                        option_ids = attr_data.get('_option_ids', [])
                        for idx, val in enumerate(attr_data.get('possible_values', [])):
                            opt_code = option_ids[idx] if idx < len(option_ids) else val[:100]
                            MarketplaceAttributeOption.objects.get_or_create(
                                attribute=ma,
                                external_code=str(opt_code)[:100],
                                defaults={'name': val[:500]},
                            )
                    saved_count += 1

                logger.info(f"Saved {saved_count} attributes to set {attr_set.id}")

            except Exception as e:
                logger.error(f"Save attributes error: {e}")

        return Response({
            'success': True,
            'attributes': all_attributes,
            'size_grid': all_size_grids if all_size_grids else None,
            'saved_count': saved_count,
            'stats': {
                'attributes_count': len(all_attributes),
                'size_grid_count': len(all_size_grids),
                'sheets_parsed': [s['sheet'] for s in sheets_to_parse if s.get('type') != 'skip'],
            }
        })

    @action(detail=False, methods=['post'])
    def save_parsed_attributes(self, request):
        """
        Сохранить распарсенные атрибуты в БД

        POST /api/ai/save_parsed_attributes/
        Body: {
            "marketplace_id": 1,
            "category_code": "4637175",
            "category_name": "Женские блузы",
            "attributes": [...]
        }
        """
        from apps.marketplaces.models import (
            Marketplace,
            MarketplaceAttributeSet,
            MarketplaceAttribute,
            MarketplaceAttributeOption
        )
        import re

        marketplace_id = request.data.get('marketplace_id')
        category_code = request.data.get('category_code')
        category_name = request.data.get('category_name', category_code)
        attributes = request.data.get('attributes', [])

        if not marketplace_id or not category_code:
            return Response(
                {'error': 'marketplace_id и category_code обязательны'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            marketplace = Marketplace.objects.get(id=marketplace_id)
        except Marketplace.DoesNotExist:
            return Response(
                {'error': 'Маркетплейс не найден'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Создаём набор атрибутов
        attr_set, set_created = MarketplaceAttributeSet.objects.get_or_create(
            marketplace=marketplace,
            external_code=category_code,
            defaults={
                'name': category_name,
                'name_uk': category_name,
            }
        )

        def slugify(text):
            if not text:
                return ''
            text = text.lower().strip()
            text = re.sub(r'[^\w\s-]', '', text)
            text = re.sub(r'[-\s]+', '_', text)
            return text[:50]

        created_attrs = []
        for attr_data in attributes:
            attr_code = attr_data.get('code') or slugify(attr_data.get('name', ''))
            if not attr_code:
                continue

            # Собираем extra_data
            extra_data = {}
            if attr_data.get('original_type'):
                extra_data['original_type'] = attr_data['original_type']
            if attr_data.get('unit'):
                extra_data['unit'] = attr_data['unit']

            attr_name = attr_data.get('name', attr_code)
            attr, attr_created = MarketplaceAttribute.objects.update_or_create(
                attribute_set=attr_set,
                external_code=attr_code,
                defaults={
                    'name': attr_name,
                    'name_uk': attr_data.get('name_uk') or attr_name,  # fallback to name
                    'attr_type': attr_data.get('type', 'string'),
                    'is_required': attr_data.get('is_required', False),
                    'is_system': False,
                    'suffix': attr_data.get('unit') or '',
                    'extra_data': extra_data,
                }
            )

            # Создаём опции
            possible_values = attr_data.get('possible_values', [])
            if possible_values and attr.attr_type in ['select', 'multiselect']:
                for i, value in enumerate(possible_values[:100]):
                    MarketplaceAttributeOption.objects.get_or_create(
                        attribute=attr,
                        external_code=slugify(str(value)) or f"opt_{i}",
                        defaults={'name': str(value)}
                    )

            created_attrs.append({
                'code': attr_code,
                'name': attr.name,
                'created': attr_created,
                'options_count': len(possible_values)
            })

        return Response({
            'success': True,
            'attribute_set_id': attr_set.id,
            'attribute_set_created': set_created,
            'attributes_count': len(created_attrs),
            'attributes': created_attrs
        })

    @action(detail=False, methods=['post'])
    def discover_attributes(self, request):
        """
        Найти атрибуты категории маркетплейса в интернете

        POST /api/ai/discover_attributes/
        Body: {
            "marketplace_name": "Rozetka",
            "category_name": "Женские блузы"
        }
        """
        from apps.marketplaces.services.ai_mapping_agent import AIWebSearchAgent

        marketplace_name = request.data.get('marketplace_name')
        category_name = request.data.get('category_name')

        if not marketplace_name or not category_name:
            return Response(
                {'error': 'marketplace_name and category_name are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            agent = AIWebSearchAgent()
            result = agent.discover_category_attributes(marketplace_name, category_name)
            return Response(result)
        except ImportError as e:
            return Response(
                {'error': 'AI agent not available: ' + str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            logger.error(f"AI discover_attributes error: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def create_discovered_attributes(self, request):
        """
        Создать атрибуты из результатов discovery в БД

        POST /api/ai/create_discovered_attributes/
        Body: {
            "marketplace_id": 1,
            "category_code": "4637175",
            "discovery_result": { ... результат discover_attributes ... }
        }
        """
        from apps.marketplaces.models import Marketplace
        from apps.marketplaces.services.ai_mapping_agent import AIWebSearchAgent

        marketplace_id = request.data.get('marketplace_id')
        category_code = request.data.get('category_code')
        discovery_result = request.data.get('discovery_result')

        if not all([marketplace_id, category_code, discovery_result]):
            return Response(
                {'error': 'marketplace_id, category_code, and discovery_result are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            marketplace = Marketplace.objects.get(id=marketplace_id)
        except Marketplace.DoesNotExist:
            return Response(
                {'error': 'Marketplace not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            agent = AIWebSearchAgent()
            result = agent.create_attributes_from_discovery(
                marketplace, category_code, discovery_result
            )
            return Response(result)
        except Exception as e:
            logger.error(f"AI create_discovered_attributes error: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def suggest_category_mappings(self, request):
        """
        AI предлагает маппинг между нашими категориями и категориями маркетплейса

        POST /api/ai/suggest_category_mappings/
        Body: {
            "marketplace_id": 1
        }
        """
        import anthropic
        from apps.marketplaces.models import Marketplace, MarketplaceCategory, CategoryMapping
        from apps.categories.models import Category

        marketplace_id = request.data.get('marketplace_id')

        if not marketplace_id:
            return Response(
                {'error': 'marketplace_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            marketplace = Marketplace.objects.get(id=marketplace_id)
        except Marketplace.DoesNotExist:
            return Response(
                {'error': 'Marketplace not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Получаем наши категории которые ещё не замаплены
        existing_mappings = CategoryMapping.objects.filter(
            marketplace_category__marketplace=marketplace
        ).values_list('category_id', flat=True)
        our_categories = Category.objects.exclude(id__in=existing_mappings).filter(
            children__isnull=True  # только листовые
        ).values('id', 'name')

        if not our_categories:
            return Response({
                'success': True,
                'suggestions': [],
                'message': 'Все категории уже замаплены'
            })

        # Получаем листовые категории маркетплейса
        mp_categories = MarketplaceCategory.objects.filter(
            marketplace=marketplace,
            children__isnull=True
        ).values('id', 'name', 'external_code')

        if not mp_categories:
            return Response({
                'success': False,
                'error': 'Нет категорий маркетплейса для маппинга'
            })

        # Формируем данные для AI
        our_cats_text = '\n'.join([f"ID:{c['id']} - {c['name']}" for c in our_categories[:50]])
        mp_cats_text = '\n'.join([f"ID:{c['id']} - {c['name']}" for c in mp_categories[:100]])

        try:
            client = anthropic.Anthropic(
                api_key=getattr(settings, 'ANTHROPIC_API_KEY', None)
            )

            prompt = f"""Ты помогаешь сопоставить категории интернет-магазина одежды с категориями маркетплейса.

НАШИ КАТЕГОРИИ (интернет-магазин одежды):
{our_cats_text}

КАТЕГОРИИ МАРКЕТПЛЕЙСА ({marketplace.name}):
{mp_cats_text}

Задача: для каждой нашей категории найди наиболее подходящую категорию маркетплейса.
Учитывай семантическое сходство названий, синонимы, варианты написания (укр/рус).

Ответь ТОЛЬКО валидным JSON массивом:
[
  {{"our_id": 123, "mp_id": 456, "confidence": 0.95, "reason": "Точное совпадение"}},
  {{"our_id": 124, "mp_id": 457, "confidence": 0.8, "reason": "Синоним: блузы = блузки"}}
]

Правила:
- confidence от 0.0 до 1.0 (уверенность в сопоставлении)
- Включай только уверенные сопоставления (confidence >= 0.6)
- Если не нашёл подходящую категорию - не включай в ответ
- reason должен быть кратким (до 50 символов)"""

            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=2000,
                messages=[{"role": "user", "content": prompt}]
            )

            result_text = response.content[0].text.strip()

            # Извлекаем JSON
            if "```json" in result_text:
                result_text = result_text.split("```json")[1].split("```")[0].strip()
            elif "```" in result_text:
                result_text = result_text.split("```")[1].split("```")[0].strip()

            suggestions = json.loads(result_text)
            logger.info(f"AI suggested {len(suggestions)} category mappings")

            # Добавляем имена категорий для UI
            our_cats_dict = {c['id']: c['name'] for c in our_categories}
            mp_cats_dict = {c['id']: c['name'] for c in mp_categories}

            for s in suggestions:
                s['our_name'] = our_cats_dict.get(s['our_id'], 'Unknown')
                s['mp_name'] = mp_cats_dict.get(s['mp_id'], 'Unknown')

            return Response({
                'success': True,
                'suggestions': suggestions,
                'total_our_categories': len(our_categories),
                'total_mp_categories': len(mp_categories)
            })

        except Exception as e:
            logger.error(f"AI suggest_category_mappings error: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def auto_discover_and_create(self, request):
        """
        Одним запросом найти и создать атрибуты для категории

        POST /api/ai/auto_discover_and_create/
        Body: {
            "marketplace_id": 1,
            "category_id": 123
        }
        """
        from apps.marketplaces.models import Marketplace, MarketplaceCategory
        from apps.marketplaces.services.ai_mapping_agent import AIWebSearchAgent

        marketplace_id = request.data.get('marketplace_id')
        category_id = request.data.get('category_id')

        if not marketplace_id or not category_id:
            return Response(
                {'error': 'marketplace_id and category_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            marketplace = Marketplace.objects.get(id=marketplace_id)
            category = MarketplaceCategory.objects.get(id=category_id, marketplace=marketplace)
        except (Marketplace.DoesNotExist, MarketplaceCategory.DoesNotExist) as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            agent = AIWebSearchAgent()

            # 1. Найти атрибуты в интернете
            discovery_result = agent.discover_category_attributes(
                marketplace.name,
                category.name
            )

            if not discovery_result.get('success'):
                return Response({
                    'success': False,
                    'phase': 'discovery',
                    'error': discovery_result.get('error')
                })

            # 2. Создать атрибуты в БД
            create_result = agent.create_attributes_from_discovery(
                marketplace,
                category.external_code,
                discovery_result
            )

            return Response({
                'success': create_result.get('success'),
                'discovery': discovery_result,
                'created': create_result
            })

        except Exception as e:
            logger.error(f"AI auto_discover_and_create error: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
