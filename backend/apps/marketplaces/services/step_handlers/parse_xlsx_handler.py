"""
Handler for parsing XLSX/Excel files with marketplace attributes.

Supports multiple formats:
- attributes_as_rows: Each row is an attribute (Epicentr style)
- attributes_as_columns: Each column is an attribute (ModnaKasta style)
- mixed: Combination of both
"""
import logging
from typing import Any
from .base import BaseStepHandler

logger = logging.getLogger(__name__)


class ParseXLSXHandler(BaseStepHandler):
    """Handler for parsing XLSX/Excel files with attributes"""

    def validate_config(self, config: dict) -> list[str]:
        errors = []
        if not config.get('filepath') and not config.get('use_previous'):
            errors.append("'filepath' is required or 'use_previous' must be true")
        return errors

    def execute(self, config: dict) -> dict:
        """
        Parse XLSX file based on AI-discovered structure.

        Config options:
            filepath: Path to XLSX file
            use_previous: Use filepath from previous step result

            # From AI findings (excel_structure):
            sheet_name: Name of sheet to parse (or list of sheet names)
            format_type: "attributes_as_rows" | "attributes_as_columns" | "mixed"
            header_row: Row number with headers (1-indexed, default: 1)

            # For attributes_as_rows format:
            attribute_id_column: Column with attribute ID (e.g., "A" or "ID")
            attribute_name_column: Column with attribute name
            attribute_type_column: Column with attribute type (optional)
            category_id_column: Column with category ID (optional)
            options_column: Column with option values (optional)
            options_separator: Separator for options (e.g., ", " or ";")

            # For attributes_as_columns format:
            skip_columns: List of columns to skip (e.g., ["A", "B"] for ID/category)
            values_start_row: Row where values start (default: 2)

            # General:
            limit: Max rows to parse (default: 10000)
            encoding: File encoding if needed

        Returns:
            {
                'sheets': list of sheet names in file,
                'parsed_sheet': name of parsed sheet,
                'format_type': detected/used format,
                'attributes': list of parsed attributes,
                'attributes_count': number of attributes found,
                'sample_data': sample of parsed data
            }
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for XLSX parsing. Install with: pip install openpyxl")

        config = self.resolve_config(config)

        filepath = config.get('filepath')
        if not filepath:
            raise ValueError("No filepath specified")

        self.log_info(f"Parsing XLSX from {filepath}")

        # Load workbook
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        result = {
            'sheets': sheet_names,
            'sheets_count': len(sheet_names),
        }

        # Determine which sheet to parse
        # Support both 'sheet_name' and legacy 'attribute_sheet' from AI findings
        sheet_name = config.get('sheet_name') or config.get('attribute_sheet')

        # Also check nested parsing_instructions from AI
        parsing_instructions = config.get('parsing_instructions', {})
        if not sheet_name:
            sheet_name = parsing_instructions.get('sheet_name') or parsing_instructions.get('attribute_sheet')

        # Merge parsing_instructions into config
        merged_config = {**config, **parsing_instructions}

        if isinstance(sheet_name, list):
            # Parse multiple sheets
            all_attributes = []
            for sn in sheet_name:
                if sn in sheet_names:
                    attrs = self._parse_sheet(wb[sn], merged_config)
                    for attr in attrs:
                        attr['source_sheet'] = sn
                    all_attributes.extend(attrs)
                else:
                    self.log_info(f"Sheet '{sn}' not found, skipping")
            result['attributes'] = all_attributes
            result['parsed_sheets'] = sheet_name
        elif sheet_name:
            if sheet_name not in sheet_names:
                # Try case-insensitive match
                matching = [s for s in sheet_names if s.lower() == sheet_name.lower()]
                if matching:
                    sheet_name = matching[0]
                else:
                    raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {sheet_names}")
            ws = wb[sheet_name]
            result['attributes'] = self._parse_sheet(ws, merged_config)
            result['parsed_sheet'] = sheet_name
        else:
            # Parse first sheet by default
            ws = wb.active
            result['attributes'] = self._parse_sheet(ws, merged_config)
            result['parsed_sheet'] = ws.title
            self.log_info(f"No sheet specified, using first sheet: {ws.title}")

        result['attributes_count'] = len(result['attributes'])
        result['format_type'] = config.get('format_type', 'auto')

        # Add sample data
        if result['attributes']:
            result['sample_data'] = result['attributes'][:5]

        wb.close()

        self.log_info(f"Parsed XLSX: {result['attributes_count']} attributes from {result.get('parsed_sheet', result.get('parsed_sheets'))}")
        return result

    def _parse_sheet(self, ws, config: dict) -> list[dict]:
        """Parse a single worksheet"""
        format_type = config.get('format_type', 'auto')
        # Also support 'format' field from AI findings
        format_field = config.get('format', '')

        if format_type == 'attributes_as_columns' or format_field == 'columns':
            self.log_info(f"Parsing sheet '{ws.title}' as columns format")
            return self._parse_columns_format(ws, config)
        elif format_type == 'attributes_as_rows' or format_field == 'rows':
            self.log_info(f"Parsing sheet '{ws.title}' as rows format")
            return self._parse_rows_format(ws, config)
        else:
            # Auto-detect based on config
            if config.get('attribute_name_column') or config.get('attribute_id_column'):
                self.log_info(f"Auto-detected rows format for sheet '{ws.title}'")
                return self._parse_rows_format(ws, config)
            else:
                self.log_info(f"Auto-detected columns format for sheet '{ws.title}'")
                return self._parse_columns_format(ws, config)

    def _parse_rows_format(self, ws, config: dict) -> list[dict]:
        """
        Parse format where each row is an attribute.
        Example:
        | ID | Name | Type | Category |
        | 51 | Brand | select | 123 |
        """
        attributes = []
        header_row = config.get('header_row', 1)
        # Support both 'data_starts_row' from AI and calculate from header_row
        data_starts_row = config.get('data_starts_row') or (header_row + 1)
        limit = config.get('limit', 10000)

        # Get column mappings
        id_col = self._col_to_index(config.get('attribute_id_column'))
        name_col = self._col_to_index(config.get('attribute_name_column'))
        type_col = self._col_to_index(config.get('attribute_type_column'))
        category_col = self._col_to_index(config.get('category_id_column'))
        options_col = self._col_to_index(config.get('options_column'))
        options_sep = config.get('options_separator', ',')

        self.log_info(f"Column mappings - ID: {id_col}, Name: {name_col}, Type: {type_col}, Category: {category_col}")

        # Read headers
        headers = []
        for cell in ws[header_row]:
            headers.append(str(cell.value) if cell.value else '')

        self.log_info(f"Headers: {headers[:10]}...")

        # Parse rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=data_starts_row, max_row=data_starts_row + limit - 1), start=1):
            cells = [cell.value for cell in row]

            # Skip empty rows
            if not any(cells):
                continue

            attr = {
                'row': row_idx + data_starts_row - 1,
                'external_id': str(cells[id_col]) if id_col is not None and id_col < len(cells) and cells[id_col] else None,
                'name': str(cells[name_col]) if name_col is not None and name_col < len(cells) and cells[name_col] else None,
            }

            if type_col is not None and type_col < len(cells):
                attr['type'] = str(cells[type_col]) if cells[type_col] else 'string'

            if category_col is not None and category_col < len(cells):
                attr['category_id'] = str(cells[category_col]) if cells[category_col] else None

            if options_col is not None and options_col < len(cells) and cells[options_col]:
                options_str = str(cells[options_col])
                attr['options'] = [o.strip() for o in options_str.split(options_sep) if o.strip()]

            # Skip if no name
            if not attr.get('name'):
                continue

            attributes.append(attr)

        self.log_info(f"Parsed {len(attributes)} attributes from rows")
        return attributes

    def _parse_columns_format(self, ws, config: dict) -> list[dict]:
        """
        Parse format where each column is an attribute (ModnaKasta style).
        Example:
        | Size | Color | Material |
        | S    | Red   | Cotton   |
        | M    | Blue  | Silk     |
        """
        attributes = []
        header_row = config.get('header_row', 1)
        # Support both 'values_start_row' and 'data_starts_row' from AI findings
        values_start_row = config.get('values_start_row') or config.get('data_starts_row') or (header_row + 1)
        skip_columns = config.get('skip_columns', [])
        limit = config.get('limit', 10000)

        # Convert skip_columns to indices (supports both ["A", "B"] and [0, 1])
        skip_indices = set()
        for col in skip_columns:
            idx = self._col_to_index(col)
            if idx is not None:
                skip_indices.add(idx)

        # Read headers (attribute names) from attribute_name_row or header_row
        attr_name_row = config.get('attribute_name_row') or header_row
        headers = []
        for cell in ws[attr_name_row]:
            headers.append(str(cell.value).strip() if cell.value else '')

        self.log_info(f"Found {len(headers)} columns, headers: {headers[:10]}...")

        # For each column, collect unique values
        for col_idx, header in enumerate(headers):
            if col_idx in skip_indices:
                continue
            if not header:
                continue

            # Collect values from this column
            options = set()
            for row_idx, row in enumerate(ws.iter_rows(min_row=values_start_row, max_row=values_start_row + limit - 1)):
                if col_idx < len(row):
                    cell_value = row[col_idx].value
                    if cell_value:
                        value = str(cell_value).strip()
                        if value:
                            options.add(value)

            attr = {
                'column': self._index_to_col(col_idx),
                'column_index': col_idx,
                'name': header,
                'type': 'select' if options else 'string',
                'options': sorted(list(options)) if options else [],
                'options_count': len(options),
            }
            attributes.append(attr)

        self.log_info(f"Parsed {len(attributes)} attributes from columns")
        return attributes

    def _col_to_index(self, col) -> int | None:
        """Convert column letter or name to 0-based index"""
        if col is None:
            return None
        if isinstance(col, int):
            return col
        if isinstance(col, str):
            col = col.strip().upper()
            if len(col) == 1 and col.isalpha():
                return ord(col) - ord('A')
            elif len(col) == 2 and col.isalpha():
                return (ord(col[0]) - ord('A') + 1) * 26 + (ord(col[1]) - ord('A'))
        return None

    def _index_to_col(self, index: int) -> str:
        """Convert 0-based index to column letter"""
        result = ""
        while index >= 0:
            result = chr(index % 26 + ord('A')) + result
            index = index // 26 - 1
        return result
