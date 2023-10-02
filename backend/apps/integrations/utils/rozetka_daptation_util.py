import os

import openpyxl
from django.db.models import Q

from apps.product.models import Product, Variant


def get_columns_numbers(sheet):
    columns = {
        'id': None,
        'rozetka_id': None,
        'name': None,
        'brand': None,
    }
    for i in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=i).value
        if value == 'Rozetka ID':
            columns['rozetka_id'] = i
        if value == 'ID из прайса продавца':
            columns['id'] = i
        elif value == 'Название товара':
            columns['name'] = i
        elif value == 'Производитель':
            columns['brand'] = i
    return columns


def get_product_pure_name(item_name, item_brand):
    item_name_lower = item_name.lower()
    brand_first_word_lower = item_brand.split(' ')[0].lower()
    pure_name = item_name_lower.split(brand_first_word_lower)[0]
    # Prettify
    pure_name_prettier = pure_name.replace(' ,', ',').replace('  ', ' ').replace('.', '').strip()
    # Capitalize first letter
    pure_name_capitalized = pure_name_prettier[0].upper() + pure_name_prettier[1:]
    return pure_name_capitalized


def loop_rows(sheet, columns):
    for i in range(2, sheet.max_row + 1):
        rozretka_id = sheet.cell(row=i, column=columns['rozetka_id']).value
        if not rozretka_id:
            continue

        size_id = sheet.cell(row=i, column=columns['id']).value
        if type(size_id) != str:
            continue
        if 'One size' in size_id:
            size_id = size_id.replace('One size', 'One_size')

        name = sheet.cell(row=i, column=columns['name']).value
        brand = sheet.cell(row=i, column=columns['brand']).value

        pure_name = get_product_pure_name(name, brand)

        product = None
        split_id = size_id.split(' ')

        if len(split_id) == 1:
            split_id = size_id.split('-')
            # join but without last element
            product = Product.objects.filter(variants__code__iexact='-'.join(split_id[:-1])).first()
            if product:
                variant = Variant.objects.filter(code__iexact='-'.join(split_id[:-1])).first()
                if variant and variant.rozetka_code != size_id:
                    variant.rozetka_code = '-'.join(split_id[:-1])
                    variant.save()
                    print(variant.rozetka_code)

        elif len(split_id) == 2:
            product = Product.objects.filter(variants__code__iexact=split_id[0]).first()

        elif len(split_id) > 2:
            rozetka_code = ' '.join(split_id[:2])
            product = Product.objects.filter(
                Q(variants__code__iexact=rozetka_code) |
                Q(variants__code__iexact='-'.join(split_id[:2]))
            ).first()

            if product:
                variant = Variant.objects.filter(code__iexact='-'.join(split_id[:2])).first()
                if variant and variant.rozetka_code != rozetka_code:
                    variant.rozetka_code = rozetka_code
                    variant.save()


        if product and product.rozetka_name != pure_name:
            product.rozetka_name = pure_name
            product.save()

        else:
            continue


def rozetka_adaptation_util(table):
    ext = os.path.splitext(table.name)[1]
    if ext.lower() == '.xlsx':
        # open file
        wb = openpyxl.load_workbook(table)

        # get active sheet
        sheet = wb.active

        # get columns names
        columns = get_columns_numbers(sheet)

        # loop rows
        loop_rows(sheet, columns)
