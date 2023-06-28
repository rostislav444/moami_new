import os

from django.core.exceptions import ValidationError
from django.db import models
from mptt.models import MPTTModel, TreeForeignKey
from openpyxl import load_workbook

from apps.abstract.models import NameSlug
from apps.sizes.models import SizeGrid


class RozetkaCategories(NameSlug, MPTTModel):
    parent = TreeForeignKey('self', null=True, blank=True, on_delete=models.CASCADE, related_name='children')
    rozetka_id = models.PositiveIntegerField()

    class MPTTMeta:
        order_insertion_by = ['name']

    class Meta:
        verbose_name = 'Категория Rozetka'
        verbose_name_plural = 'Категории Rozetka'

    def __str__(self):
        categories = self.get_ancestors(include_self=True)
        if categories.count():
            return ' / '.join([c.name for c in categories])
        return self.name


class GoogleTaxonomy(models.Model):
    id = models.PositiveIntegerField(primary_key=True)
    name = models.CharField(max_length=1000)
    name_ru = models.CharField(max_length=1000)

    class Meta:
        ordering = ['name']
        verbose_name = 'Категории: категория Google Taxonomy'
        verbose_name_plural = 'Категории: категории Google Taxonomy'

    def __str__(self):
        if self.name_ru:
            return ' - '.join([str(self.id), self.name_ru])
        else:
            return ' - '.join([str(self.id), self.name])

    def save(self):
        super(GoogleTaxonomy, self).save()


def validate_xlsx(value):
    ext = os.path.splitext(value.name)[1]  # [0] returns path+filename
    valid_extensions = ['.xlsx']
    if not ext.lower() in valid_extensions:
        raise ValidationError(u'Загрузите таблицу EXCEL в формате .xlsx')


def validate_rar(value):
    ext = os.path.splitext(value.name)[1]  # [0] returns path+filename
    valid_extensions = ['.rar']
    if not ext.lower() in valid_extensions:
        raise ValidationError(u'Загрузите архив в формате .rar')


class GoogleTaxonomyUplaoder(models.Model):
    table = models.FileField(upload_to='google_taxonomy', validators=[validate_xlsx], help_text='.xlsx')
    table_ru = models.FileField(upload_to='google_taxonomy', validators=[validate_xlsx], help_text='.xlsx')
    delate_all = models.BooleanField(default=False)

    class Meta:
        verbose_name = 'Загрузчик: категория Google Taxonomy'
        verbose_name_plural = 'Загрузчик: категории Google Taxonomy'

    def save(self):
        if self.delate_all == True:
            GoogleTaxonomy.objects.all().delete()

        super(GoogleTaxonomyUplaoder, self).save()
        # OPEN TABLE EN
        xlsx_table = load_workbook(self.table)
        sheets = xlsx_table.sheetnames
        sheet = xlsx_table[sheets[0]]
        # ITER
        for num, row in enumerate(sheet.values):
            name = ''
            columns = [col for col in row if col != None]
            id = columns[0]
            columns = columns[1:]
            for n, column in enumerate(columns):
                name += column
                if n < len(columns) - 1:
                    name += ' > '
            # SAVE
            try:
                taxonomy = GoogleTaxonomy.objects.get(id=id)
            except:
                taxonomy = GoogleTaxonomy(id=id, name=name)
            taxonomy.save()
        xlsx_table.close()

        # OPEN TABLE RU
        xlsx_table_ru = load_workbook(self.table_ru)
        sheets_ru = xlsx_table_ru.sheetnames
        sheet_ru = xlsx_table_ru[sheets_ru[0]]
        # ITER
        for num, row in enumerate(sheet_ru.values):
            name = ''
            columns = [col for col in row if col != None]
            print(columns)
            id = columns[0]
            columns = columns[1:]
            for n, column in enumerate(columns):
                name += column
                if n < len(columns) - 1:
                    name += ' > '
            # SAVE
            try:
                taxonomy = GoogleTaxonomy.objects.get(id=id)
                taxonomy.name_ru = name
                taxonomy.save()
            except:
                pass
        xlsx_table_ru.close()
